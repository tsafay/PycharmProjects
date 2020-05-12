#！python3
# -*- coding: utf-8 -*-
# @Author  : lee 
# @Date    : 2020/5/5 14:34
'项目：电子表格单元格转置程序'
import openpyxl
wb = openpyxl.load_workbook('news.xlsx')
sheet = wb['news']
print(list(sheet.columns))
print(list(sheet.rows))
l = list(sheet.columns)
print(l[0][0])
wb1 = openpyxl.Workbook()
sheet1 = wb1.active
sheet1.title = 'nmm'

for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        sheet1.cell(row=j,column=i).value = sheet.cell(row=i,column=j).value

#wb1.save('news3.xlsx')




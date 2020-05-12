#！python3
# -*- coding: utf-8 -*-
# @Author  : lee 
# @Date    : 2020/5/4 21:50
'写入excel文档'
import openpyxl
wb = openpyxl.Workbook()  #创建新的空workbook对象
# 工作簿从一个工作表开始，默认名为sheet，可以将新的字符串保存在title属性中改名名称
ws = wb.active
ws.title = 'ooop'
# 创建新的工作表
wb.create_sheet(index=0,title='news')
wb.create_sheet(index=1,title='news2')
wb.create_sheet(index=2,title='news3')

wb.save('news.xlsx')  # 保存表格

# 删除sheet表的两种方法
del wb['news3']
wb.remove(wb['news2'])

sheet = wb['ooop']  # 获取sheet表名称
sheet['A1'] = 'hello' # 将值写入单元格
wb.save('news2.xlsx')





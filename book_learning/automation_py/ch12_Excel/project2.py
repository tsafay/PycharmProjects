#！python3
# -*- coding: utf-8 -*-
# @Author  : lee 
# @Date    : 2020/5/5 13:44
'项目：空行插入程序'
'''
程序从表格第n行开始，插入m个空行；
'''
import openpyxl
row_num = int(input('行号：'))
enter_num = int(input('插入空行：'))
filename = input('表格名称：')
wb = openpyxl.load_workbook(filename)
sheet = wb.active

wb1 = openpyxl.Workbook()
sheet1 = wb1.active
sheet1.title = 'news'

for i in range(1,row_num):
    for j in range(1,sheet.max_column+1):
        sheet1.cell(row=i,column=j).value = sheet.cell(row=i,column=j).value

for x in range(row_num,sheet.max_row+1):
    for y in range(1,sheet.max_column+1):
        sheet1.cell(row=x+enter_num,column=y).value = sheet.cell(row=x,column=y).value

wb1.save('new_'+filename)
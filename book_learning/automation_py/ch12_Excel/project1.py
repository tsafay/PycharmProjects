#！python3
# -*- coding: utf-8 -*-
# @Author  : lee 
# @Date    : 2020/5/5 10:05
'实践项目：乘法表'
'''
创建程序，从命令行接受参数N，在excel电子表格中创建一个N*N的乘法表；
行1和列A应该用作标签，使用粗体。
'''
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

num = int(input('输入参数：'))
wb = openpyxl.Workbook()

sheet = wb['Sheet']
sheet.title = 'Multiplication table'
sheet['A1'] = 'table'
for i in range(2,num+2):
    sheet['A'+str(i)] = i-1
    sheet['A' + str(i)].font = Font(size=24,bold=True)
    sheet[get_column_letter(i)+str(1)] = i-1
    sheet[get_column_letter(i) + str(1)].font = Font(bold=True)

    for x in range(2,num+2):
        sheet[get_column_letter(x)+str(i)] = (x-1)*(i-1)

wb.save('mult.xlsx')



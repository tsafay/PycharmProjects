#！python3
# -*- coding: utf-8 -*-
# @Author  : lee 
# @Date    : 2020/5/5 20:55
'项目：excel到csv文件转换'
import openpyxl,csv,os
os.makedirs('csv_ex',exist_ok=True)
for file in os.listdir('.'):
    if not file.endswith('.xlsx'):
        continue
    file_work = openpyxl.load_workbook(file)

    for sheet in file_work.sheetnames:
        sheetdata = []
        for row in range(1,file_work[sheet].max_row+1):
            row_data = []
            for col in range(1,file_work[sheet].max_column+1):
                row_data.append(file_work[sheet].cell(row=row,column=col).value)
            sheetdata.append(row_data)

            csvfile = open(os.path.join('csv_ex',sheet +'.csv'),
                           'w',newline='')
            csvread = csv.writer(csvfile)
            for row in sheetdata:
                csvread.writerow(row)
            csvfile.close()


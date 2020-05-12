#！python3
# -*- coding: utf-8 -*-
# @Author  : lee 
# @Date    : 2020/5/6 22:08
'项目：向会员发送会费提醒电子邮件'
'''
程序任务：
 从 Excel 电子表格中读取数据。
 找出上个月没有交费的所有会员。
 找到他们的电子邮件地址，向他们发送针对个人的提醒。
程序实现：
 用 openpyxl 模块打开并读取 Excel 文档的单元格。
 创建一个字典，包含会费超期的会员。
 调用 smtplib.SMTP()、ehlo()、starttls()和 login()，登录 SMTP 服务器。
 针对会费超期的所有会员，调用 sendmail()方法，发送针对个人的电子邮件提醒。
'''
import openpyxl,sys,smtplib

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.active

lastCol = sheet.max_column
latestMonth = sheet.cell(row=1,column=lastCol).value

# TODO: 检查每个会员的付款状态
unpaidMembers = {}
for i in range(2,sheet.max_row+1):
    payment = sheet.cell(row=i,column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=i,column=1).value
        email = sheet.cell(row=i,column=2).value
        unpaidMembers[name] = email

# TODO: 登录到邮件账户
captcha = input('输入邮箱验证码：')
smt = smtplib.SMTP('smtp.qq.com',587)
smt.ehlo()
smt.starttls()
smt.login('287672785@qq.com',captcha)

# TODO: 发送邮件
for name,email in unpaidMembers.items():
    body = 'Sbuject: %s dues unpaid.\nDear %s,\nRecords show that you have ' \
           'not paid dues for %s.Please make this payment as soon as possible.' \
           'Thank you!' %(latestMonth,name,latestMonth)
    print('Sending email to %s...' % email)
    sendmailStatus = smt.sendmail('287672785@qq.com',email,body)

    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s.'
              %(email,sendmailStatus))
smt.quit()